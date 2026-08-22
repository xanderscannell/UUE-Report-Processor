#!/usr/bin/env python3
"""
Daily Events Excel Reader
=========================
Reads the events database's "Daily Events - Excel" export and produces the same
event records the PDF parser does, so everything downstream — location
filtering, schedule rows, sorting, output, and the Gantt feed — is shared.

The export has two sheets:

* ``Parameter Summary`` — label/value pairs, including the report date
* ``Event List <date>`` — one row per **booking**, headers on the first row

A booking is an event in one room, which is already how the PDF path models
things, so an event occupying two rooms correctly yields two records.

Note the export carries no setup-start time: ``Setup Ready By`` comes from
``Event Start``, matching the PDF parser's own third fallback. See ADR-008.
"""

import re
import logging
import warnings
from contextlib import contextmanager
from datetime import datetime
from typing import Dict, List, Optional

from openpyxl import load_workbook

from setup_report_processor import EventScheduleProcessor

# A child of the logger the GUI attaches its panel handler to, so this module's
# records propagate there instead of stopping at the root logger.
logger = logging.getLogger("setup_report_processor.daily_events_excel")

# Sheet holding the report's parameters (report date, search names, ...).
PARAMETER_SHEET = "Parameter Summary"

# Data sheets are named "Event List <date>", so match on the prefix.
DATA_SHEET_PREFIX = "event list"

# Label whose neighbouring cell holds the report date.
REPORT_DATE_LABEL = "report date"

# Columns the reader cannot work without.
REQUIRED_COLUMNS = ("Event Start", "Event End", "Event Name", "Location")

# Date formats seen in the export, tried in order.
DATE_FORMATS = ("%b %d %Y", "%B %d %Y", "%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d")

# Trailing date in a sheet title, e.g. "Event List Aug 22 2026".
SHEET_DATE_PATTERN = re.compile(r"([A-Za-z]{3,9}\s+\d{1,2}\s+\d{4})\s*$")


def _text(value) -> str:
    """Render a cell as a stripped string, treating None as empty."""
    return "" if value is None else str(value).strip()


def _format_time(value) -> Optional[str]:
    """
    Render a cell as the "7:30 AM" string the rest of the pipeline expects.

    Built by hand rather than with ``strftime`` because ``%I``/``%p`` are
    locale-dependent, and every downstream regex and ``parse_time()`` call
    assumes the PDF's English 12-hour format.

    Args:
        value: A datetime/time cell, or a string for exports that emit text

    Returns:
        Time string like "9:00 AM", or None if the cell holds no usable time

    Example:
        >>> _format_time(datetime(2026, 8, 22, 15, 30))
        '3:30 PM'
    """
    if isinstance(value, datetime):
        hour, minute = value.hour, value.minute
    elif hasattr(value, "hour") and hasattr(value, "minute"):
        # datetime.time, for exports that store a bare time-of-day.
        hour, minute = value.hour, value.minute
    else:
        text = _text(value)
        if not text:
            return None
        match = re.search(r"(\d{1,2}):(\d{2})\s*([AaPp])\.?[Mm]\.?", text)
        if not match:
            logger.warning(f"Could not read a time from: {text!r}")
            return None
        hour, minute = int(match.group(1)) % 12, int(match.group(2))
        if match.group(3).upper() == "P":
            hour += 12

    return f"{hour % 12 or 12}:{minute:02d} {'AM' if hour < 12 else 'PM'}"


def _parse_date(value) -> Optional[datetime]:
    """Parse a report-date cell, which may already be a datetime."""
    if isinstance(value, datetime):
        return value

    text = _text(value)
    if not text:
        return None

    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


class DailyEventsExcelProcessor(EventScheduleProcessor):
    """Process Daily Events Excel exports and extract event schedules."""

    SOURCE_SUFFIXES = (".xlsx",)
    SOURCE_LABEL = "Excel"

    def _validate_suffix(self) -> None:
        """
        Check the extension, with a pointed message for the legacy .xls format.

        Raises:
            ValueError: If the file is not a readable Excel workbook
        """
        if self.source_path.suffix.lower() == ".xls":
            raise ValueError(
                "Legacy .xls workbooks cannot be read - "
                "re-save the report as .xlsx"
            )
        super()._validate_suffix()

    @contextmanager
    def _workbook(self):
        """Open the workbook read-only, closing it on the way out."""
        with warnings.catch_warnings():
            # These exports ship a minimal styles.xml, so openpyxl warns
            # about the missing default style on every single file.
            warnings.filterwarnings(
                "ignore",
                message="Workbook contains no default style",
                category=UserWarning,
            )
            workbook = load_workbook(
                self.source_path, read_only=True, data_only=True
            )
        try:
            yield workbook
        finally:
            workbook.close()

    # ==================================================================
    # Report date
    # ==================================================================
    def extract_report_date(self) -> Optional[str]:
        """
        Extract the report date and format it as MM-DD-YY.

        Falls back in this order:
        1. The "Report Date" row of the Parameter Summary sheet
        2. The date at the end of the data sheet's title
        3. The first "Day" cell in the data sheet

        The chain is worth having because these exports carry a generic
        filename (``DailyEventsExcel.xlsx``), so the filename fallback in
        ``get_output_basename()`` is a poor last resort.

        Returns:
            Formatted date string (e.g., "08-22-26") or None if not found

        Example:
            >>> processor.extract_report_date()
            '08-22-26'
        """
        try:
            with self._workbook() as workbook:
                for source in (
                    self._date_from_parameters,
                    self._date_from_sheet_title,
                    self._date_from_day_column,
                ):
                    parsed = source(workbook)
                    if parsed:
                        return parsed.strftime("%m-%d-%y")
        except Exception as e:
            logger.warning(f"Error extracting report date: {e}")
            return None

        logger.warning("Report date not found in the workbook")
        return None

    def _date_from_parameters(self, workbook) -> Optional[datetime]:
        """Read the date from the Parameter Summary sheet's Report Date row."""
        if PARAMETER_SHEET not in workbook.sheetnames:
            return None

        for row in workbook[PARAMETER_SHEET].iter_rows(values_only=True):
            labelled = False
            for cell in row:
                if labelled:
                    parsed = _parse_date(cell)
                    if parsed:
                        return parsed
                elif _text(cell).lower().rstrip(":") == REPORT_DATE_LABEL:
                    labelled = True
        return None

    def _date_from_sheet_title(self, workbook) -> Optional[datetime]:
        """Read the date off the end of an "Event List <date>" sheet title."""
        for title in self._data_sheet_titles(workbook):
            match = SHEET_DATE_PATTERN.search(title)
            if match:
                parsed = _parse_date(match.group(1))
                if parsed:
                    return parsed
        return None

    def _date_from_day_column(self, workbook) -> Optional[datetime]:
        """Read the date from the first populated Day cell in the data sheet."""
        for title in self._data_sheet_titles(workbook):
            rows = workbook[title].iter_rows(values_only=True)
            header = next(rows, None)
            if header is None:
                continue

            index = self._header_index(header, title, require=False)
            position = index.get("day")
            if position is None:
                continue

            for values in rows:
                if position < len(values):
                    parsed = _parse_date(values[position])
                    if parsed:
                        return parsed
        return None

    # ==================================================================
    # Events
    # ==================================================================
    def _data_sheet_titles(self, workbook) -> List[str]:
        """
        Return the titles of the sheets holding event rows.

        Every "Event List ..." sheet is included, so a multi-day export is read
        whole rather than silently truncated to its first day. If none match,
        falls back to the first sheet that is not the Parameter Summary.

        Args:
            workbook: An open openpyxl workbook

        Returns:
            List of sheet titles, in workbook order
        """
        titles = [
            title for title in workbook.sheetnames
            if title.strip().lower().startswith(DATA_SHEET_PREFIX)
        ]
        if titles:
            return titles

        return [
            title for title in workbook.sheetnames
            if title.strip().lower() != PARAMETER_SHEET.lower()
        ][:1]

    def _header_index(
        self, header, sheet_title: str, require: bool = True
    ) -> Dict[str, int]:
        """
        Map lower-cased column names to their position in the row.

        Args:
            header: The sheet's first row, as a tuple of values
            sheet_title: Sheet name, used in the error message
            require: Whether to enforce the presence of REQUIRED_COLUMNS

        Returns:
            Dict of lower-cased column name to column index

        Raises:
            ValueError: If require is True and a required column is absent
        """
        index: Dict[str, int] = {}
        for position, cell in enumerate(header):
            name = _text(cell).lower()
            if name:
                index.setdefault(name, position)

        if require:
            missing = [
                column for column in REQUIRED_COLUMNS
                if column.lower() not in index
            ]
            if missing:
                # Named columns lead the message: a file card shows only the
                # first ~60 characters, and the column name is the actionable
                # part.
                raise ValueError(
                    f"Missing required column(s): {', '.join(missing)} "
                    f"- in sheet '{sheet_title}'"
                )

        return index

    def _collect_events(self) -> List[Dict[str, str]]:
        """
        Read every event row in the workbook and filter it by location.

        Returns:
            List of event dictionaries

        Raises:
            ValueError: If the workbook has no data sheet, or a data sheet is
                missing one of the required columns
        """
        logger.info("Reading events from the Daily Events export...")
        events: List[Dict[str, str]] = []
        total_rows = 0

        with self._workbook() as workbook:
            titles = self._data_sheet_titles(workbook)
            if not titles:
                raise ValueError(
                    "No event sheet found - expected a sheet named "
                    f"'{DATA_SHEET_PREFIX.title()} <date>'"
                )

            for title in titles:
                rows = workbook[title].iter_rows(values_only=True)
                header = next(rows, None)
                if header is None:
                    logger.warning(f"Sheet '{title}' is empty")
                    continue

                index = self._header_index(header, title)
                logger.debug(f"Reading sheet '{title}'")

                for values in rows:
                    if not any(_text(cell) for cell in values):
                        continue

                    total_rows += 1
                    try:
                        event = self._parse_event_row(values, index)
                    except Exception as e:
                        logger.warning(f"Error parsing event row: {e}")
                        continue

                    if event:
                        events.append(event)

        excluded_count = total_rows - len(events)
        logger.info(
            f"Found {len(events)} valid events in {self.source_path.name} "
            f"(excluded {excluded_count} events)"
        )
        return events

    def _parse_event_row(
        self, values, index: Dict[str, int]
    ) -> Optional[Dict[str, str]]:
        """
        Parse a single booking row into an event record.

        Args:
            values: The row's cell values
            index: Column name to position map from _header_index()

        Returns:
            Dictionary with event details or None if the row is excluded
        """
        def column(name: str):
            position = index.get(name.lower())
            if position is None or position >= len(values):
                return None
            return values[position]

        # Event Name is always populated; Event Title is the fuller form the
        # export sometimes carries instead.
        event_name = _text(column("Event Name")) or _text(column("Event Title"))
        if not event_name:
            logger.info("EXCLUDED: Event with no name found")
            return None

        raw_location = _text(column("Location"))
        if not raw_location:
            logger.info(f"EXCLUDED: '{event_name}' - no valid location found")
            return None

        # The export's Location codes ("UC 1225") are exactly the whitelist's,
        # so the PDF path's matcher is reused unchanged.
        location = self._match_whitelist_location(raw_location)
        if not location:
            logger.info(
                f"EXCLUDED: '{event_name}' at '{raw_location}' - "
                "location not in whitelist"
            )
            return None

        # No setup-start column exists, so the event's own start time is the
        # point the room must be ready by (ADR-008).
        setup_time = _format_time(column("Event Start"))
        closing_time = _format_time(column("Event End"))
        if not setup_time or not closing_time:
            logger.info(f"EXCLUDED: '{event_name}' - no event times found")
            return None

        return {
            "event_name": event_name,
            "location": location,
            "setup_time": setup_time,
            "closing_time": closing_time,
        }
